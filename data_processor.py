#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
data_processor.py - Універсальний інструмент для роботи з CSV та Excel файлами
==============================================================================

Функціональність:
1. Конвертація CSV ↔ XLSX (з автовизначенням кодування та роздільників)
2. Фільтрація даних (6 типів фільтрів: текст ==, містить, список, число ==, діапазон, regex)
3. Об'єднання файлів/аркушів з різними структурами
4. Зведення та аналіз (частоти, унікальні значення, групування)
5. Дедуплікація за ключовими колонками
6. Паралельна обробка великих CSV файлів

Залежності:
    pip install pandas openpyxl xlsxwriter tqdm

Автор: Об'єднання скриптів csv_worker, csv_parallel_orchestrator,
       csv_semicolon_to_xlsx, xlsx_group_summary_interactive, xlsx_unify_unique_interactive
"""

import sys
import os
import re
import csv
import json
import base64
import argparse
import tempfile
import shutil
import subprocess
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any, Callable, Set
from datetime import datetime

try:
    import pandas as pd
    import numpy as np
except ImportError:
    sys.exit("Помилка: Встановіть pandas: pip install pandas")

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("[WARN] openpyxl не встановлено. XLSX функції обмежені.")
    openpyxl = None

try:
    import xlsxwriter
    EXCEL_ENGINE = "xlsxwriter"
except ImportError:
    try:
        import openpyxl
        EXCEL_ENGINE = "openpyxl"
    except ImportError:
        EXCEL_ENGINE = None

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None


# ============================================================================
# КОНСТАНТИ
# ============================================================================

COMMON_ENCODINGS = ["utf-8", "utf-8-sig", "cp1251", "windows-1251", "latin-1", "cp1252"]
COMMON_SEPS = [",", ";", "\t", "|", ":"]
CHUNKSIZE = 200_000
MAX_EXCEL_ROWS = 1_048_576

# Мапа для нормалізації назв колонок (кирилиця → латиниця)
CYR_TO_LAT = str.maketrans({
    "А":"A", "В":"B", "С":"C", "Е":"E", "Н":"H", "К":"K", "М":"M", "О":"O",
    "Р":"P", "Т":"T", "Х":"X", "У":"Y", "І":"I", "Ї":"I", "Й":"I", "Ґ":"G",
    "а":"a", "в":"b", "с":"c", "е":"e", "н":"h", "к":"k", "м":"m", "о":"o",
    "р":"p", "т":"t", "х":"x", "у":"y", "і":"i", "ї":"i", "й":"i", "ґ":"g"
})

# Мапа типографських лапок
QUOTE_MAP = str.maketrans({
    """: '"', """: '"', "„": '"', "‟": '"', "«": '"', "»": '"',
    "‚": "'", "'": "'", "'": "'", "‹": "'", "›": "'", "´": "'", "`": "'"
})


# ============================================================================
# УТИЛІТИ ДЛЯ РОБОТИ З ФАЙЛАМИ
# ============================================================================

def detect_encoding_and_sep(path: str) -> Tuple[str, str]:
    """Автовизначення кодування та роздільника для CSV файлу"""
    head = ""
    enc_used = "utf-8"

    for enc in COMMON_ENCODINGS:
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                head = f.read(8192)
            enc_used = enc
            break
        except Exception:
            continue

    if not head:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            head = f.read(8192)
        enc_used = "utf-8"

    # Визначення роздільника
    try:
        dialect = csv.Sniffer().sniff(head, delimiters="".join(COMMON_SEPS))
        sep = dialect.delimiter
    except Exception:
        counts = {d: head.count(d) for d in COMMON_SEPS}
        sep = max(counts, key=counts.get) if counts else ","

    return enc_used, sep


def read_file_auto(path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """Універсальне читання CSV або XLSX з автовизначенням параметрів"""
    p = Path(path)

    if p.suffix.lower() == ".csv":
        enc, sep = detect_encoding_and_sep(path)
        print(f"[INFO] Виявлено: encoding={enc}, separator='{sep}'")
        return pd.read_csv(path, encoding=enc, sep=sep, engine="python",
                          on_bad_lines="warn", dtype=str)

    elif p.suffix.lower() in [".xlsx", ".xls"]:
        if sheet_name:
            return pd.read_excel(path, sheet_name=sheet_name, dtype=str)
        else:
            return pd.read_excel(path, dtype=str)

    else:
        raise ValueError(f"Непідтримуваний формат файлу: {p.suffix}")


def save_to_excel(df: pd.DataFrame, output_path: str, sheet_name: str = "Data",
                 freeze_header: bool = True, autofilter: bool = True,
                 force_text_cols: Optional[List[str]] = None):
    """Збереження DataFrame в Excel з форматуванням"""
    if not EXCEL_ENGINE:
        raise RuntimeError("Встановіть xlsxwriter або openpyxl: pip install xlsxwriter openpyxl")

    with pd.ExcelWriter(output_path, engine=EXCEL_ENGINE) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]

        if freeze_header:
            try:
                if EXCEL_ENGINE == "xlsxwriter":
                    ws.freeze_panes(1, 0)
                else:
                    ws.freeze_panes = "A2"
            except Exception:
                pass

        if autofilter:
            try:
                if EXCEL_ENGINE == "xlsxwriter":
                    ws.autofilter(0, 0, len(df), len(df.columns) - 1)
                else:
                    ws.auto_filter.ref = ws.dimensions
            except Exception:
                pass

        # Форматування текстових колонок
        if force_text_cols and EXCEL_ENGINE == "xlsxwriter":
            fmt = writer.book.add_format({"num_format": "@"})
            for col_name in force_text_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name)
                    ws.set_column(col_idx, col_idx, None, fmt)


# ============================================================================
# НОРМАЛІЗАЦІЯ ТА РЕЗОЛЮЦІЯ КОЛОНОК
# ============================================================================

def norm_name(s: str) -> str:
    """Нормалізація назви колонки (без регістру, пробілів, дефісів)"""
    if s is None:
        return ""
    t = str(s).strip().lower()
    t = t.replace("і", "i")
    t = t.translate(CYR_TO_LAT)
    t = re.sub(r"[\s\-\u00A0]+", "", t)
    return t


def norm_col_name(name: Any) -> str:
    """Повна нормалізація для зіставлення колонок"""
    s = "" if name is None else str(name)
    s = s.translate(CYR_TO_LAT)
    s = s.strip().lower()
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"[^0-9a-z_]+", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def resolve_single_column(df: pd.DataFrame, user_token: str) -> str:
    """Знаходження колонки за назвою або літерою (A, B, C...)"""
    # Спочатку точний збіг
    if user_token in df.columns:
        return user_token

    token = str(user_token).strip()

    # Спроба як літера колонки (A, B, C...)
    if re.fullmatch(r"[A-Za-z]+", token):
        try:
            idx = column_index_from_string(token.upper()) - 1
            if 0 <= idx < df.shape[1]:
                return df.columns[idx]
        except Exception:
            pass

    # Пошук по нормалізованій назві
    target = norm_name(user_token)
    for col in df.columns:
        if norm_name(col) == target:
            return col

    # Часткове співпадіння
    for col in df.columns:
        if target in norm_name(col) or norm_name(col) in target:
            return col

    raise KeyError(f"Стовпець '{user_token}' не знайдено. Доступні: {list(df.columns)}")


def resolve_multi_columns(df: pd.DataFrame, token: str) -> List[str]:
    """Резолюція множинних колонок (B:D, B,D,F або назви через кому)"""
    tok = token.strip()
    cols: List[str] = []

    # Діапазон: B:D
    if re.fullmatch(r"[A-Za-z]+:[A-Za-z]+", tok):
        a, b = tok.split(":")
        ia = column_index_from_string(a.upper()) - 1
        ib = column_index_from_string(b.upper()) - 1
        if ia > ib:
            ia, ib = ib, ia
        for j in range(max(0, ia), min(df.shape[1] - 1, ib) + 1):
            cols.append(df.columns[j])
        return cols

    # Список літер: B,D,F
    if re.fullmatch(r"[A-Za-z]+(,[A-Za-z]+)*", tok):
        for t in tok.split(","):
            idx = column_index_from_string(t.strip().upper()) - 1
            cols.append(df.columns[idx])
        return cols

    # Список назв через кому
    names = [t.strip() for t in tok.split(",") if t.strip()]
    for nm in names:
        cols.append(resolve_single_column(df, nm))

    return cols


# ============================================================================
# ФІЛЬТРИ (6 ТИПІВ)
# ============================================================================

FILTERS_HELP = """
[Довідка по фільтрам]
0) Без фільтрації
1) Текст дорівнює (точний збіг)
2) Текст містить (case-insensitive)
3) Текст у списку (isin)
4) Число дорівнює
5) Число в діапазоні [min; max]
6) REGEX (регулярний вираз Python)

Всі фільтри комбінуються через AND (логічне "І").
"""


def normalize_series(s: pd.Series, case: str = "upper", strip: bool = True) -> pd.Series:
    """Нормалізація Series для фільтрації"""
    s = s.fillna("").astype(str)
    if strip:
        s = s.str.strip()
    if case == "upper":
        s = s.str.upper()
    elif case == "lower":
        s = s.str.lower()
    return s


def build_filter_from_spec(df: pd.DataFrame, spec: Dict[str, Any]) -> Optional[pd.Series]:
    """Побудова маски фільтрації з специфікації"""
    if not spec or spec.get("mode") == "0":
        return None

    mode = str(spec.get("mode", "0"))
    col_token = spec.get("column", "")

    try:
        col = resolve_single_column(df, col_token)
    except Exception as e:
        print(f"[WARN] Фільтр пропущено ({col_token}): {e}")
        return None

    mask = pd.Series(True, index=df.index)

    # Текст дорівнює
    if mode == "1":
        val = spec.get("value", "")
        case = spec.get("case", "upper")
        strip_ws = spec.get("strip_ws", True)
        s = normalize_series(df[col], case=case, strip=strip_ws)
        if strip_ws:
            val = val.strip()
        if case == "upper":
            val = val.upper()
        elif case == "lower":
            val = val.lower()
        mask = (s == val)

    # Текст містить
    elif mode == "2":
        sub = spec.get("value", "")
        case = spec.get("case", "upper")
        s = normalize_series(df[col], case=case, strip=spec.get("strip_ws", True))
        if case == "upper":
            sub = sub.upper()
        elif case == "lower":
            sub = sub.lower()
        mask = s.str.contains(re.escape(sub), na=False)

    # Текст у списку
    elif mode == "3":
        values = spec.get("values", [])
        case = spec.get("case", "upper")
        s = normalize_series(df[col], case=case, strip=spec.get("strip_ws", True))
        if case == "upper":
            values = [x.upper() for x in values]
        elif case == "lower":
            values = [x.lower() for x in values]
        mask = s.isin(values)

    # Число дорівнює
    elif mode == "4":
        try:
            target = float(str(spec.get("value", "")).replace(",", "."))
        except Exception:
            target = None
        if target is not None:
            s = pd.to_numeric(df[col], errors="coerce")
            mask = (s == target)
        else:
            mask = pd.Series(False, index=df.index)

    # Число в діапазоні
    elif mode == "5":
        try:
            vmin = float(str(spec.get("min", "")).replace(",", "."))
        except Exception:
            vmin = float("-inf")
        try:
            vmax = float(str(spec.get("max", "")).replace(",", "."))
        except Exception:
            vmax = float("inf")
        s = pd.to_numeric(df[col], errors="coerce")
        mask = s.between(vmin, vmax, inclusive="both")

    # REGEX
    elif mode == "6":
        pattern = spec.get("pattern", ".*")
        try:
            rx = re.compile(pattern)
            s = normalize_series(df[col], case=spec.get("case", "keep"),
                               strip=spec.get("strip_ws", True))
            mask = s.map(lambda x: bool(rx.search(x)) if isinstance(x, str) else False)
        except Exception as e:
            print(f"[WARN] REGEX помилка: {e}")
            return None

    return mask


def apply_filters(df: pd.DataFrame, filters: List[Dict[str, Any]]) -> pd.DataFrame:
    """Застосування множинних фільтрів до DataFrame"""
    result = df.copy()

    for i, spec in enumerate(filters, 1):
        mask = build_filter_from_spec(result, spec)
        if mask is not None:
            result = result[mask].copy()
            print(f"[INFO] Фільтр {i}: залишилось {len(result):,} рядків")

    return result


def prompt_filters() -> List[Dict[str, Any]]:
    """Інтерактивне створення фільтрів"""
    print(FILTERS_HELP)

    try:
        n = int(input("Скільки фільтрів застосувати? (0-6) [0]: ").strip() or "0")
    except Exception:
        n = 0

    n = max(0, min(6, n))
    filters = []

    for i in range(1, n + 1):
        print(f"\n=== Фільтр {i} ===")
        mode = input("Тип (1=TXT==, 2=CONTAINS, 3=LIST, 4=NUM==, 5=RANGE, 6=REGEX) [2]: ").strip() or "2"

        spec = {"mode": mode}
        spec["column"] = input("Назва колонки [A]: ").strip() or "A"

        if mode in ("1", "2", "3", "6"):
            spec["case"] = input("Регістр (keep/upper/lower) [upper]: ").strip().lower() or "upper"
            spec["strip_ws"] = input("Обрізати пробіли? (y/n) [y]: ").strip().lower() != "n"

        if mode == "1":
            spec["value"] = input("Значення: ").strip()
            spec["force_text"] = input("Форматувати як ТЕКСТ? (y/n) [n]: ").strip().lower() == "y"

        elif mode == "2":
            spec["value"] = input("Підрядок: ").strip()
            spec["force_text"] = input("Форматувати як ТЕКСТ? (y/n) [n]: ").strip().lower() == "y"

        elif mode == "3":
            vals = input("Список через кому: ").strip()
            spec["values"] = [x.strip() for x in vals.split(",") if x.strip()]
            spec["force_text"] = input("Форматувати як ТЕКСТ? (y/n) [n]: ").strip().lower() == "y"

        elif mode == "4":
            spec["value"] = input("Числове значення: ").strip()

        elif mode == "5":
            spec["min"] = input("Мінімум: ").strip()
            spec["max"] = input("Максимум: ").strip()

        elif mode == "6":
            spec["pattern"] = input("REGEX шаблон: ").strip() or ".*"

        filters.append(spec)

    return filters


# ============================================================================
# КОНВЕРТАЦІЯ CSV ↔ XLSX
# ============================================================================

def csv_to_xlsx(input_csv: str, output_xlsx: str, encoding: Optional[str] = None,
                separator: Optional[str] = None, force_text_cols: Optional[List[str]] = None):
    """Конвертація CSV в XLSX з автовизначенням параметрів"""
    print(f"[INFO] Конвертація {input_csv} → {output_xlsx}")

    # Автовизначення параметрів якщо не задані
    if encoding is None or separator is None:
        enc, sep = detect_encoding_and_sep(input_csv)
        encoding = encoding or enc
        separator = separator or sep

    print(f"[INFO] Encoding: {encoding}, Separator: '{separator}'")

    # Читання CSV
    df = pd.read_csv(input_csv, encoding=encoding, sep=separator,
                     engine="python", on_bad_lines="warn", dtype=str)

    print(f"[INFO] Завантажено {len(df):,} рядків, {len(df.columns)} колонок")

    # Збереження в XLSX
    save_to_excel(df, output_xlsx, sheet_name="Data",
                 force_text_cols=force_text_cols)

    print(f"[OK] Створено: {output_xlsx}")


def xlsx_to_csv(input_xlsx: str, output_csv: str, sheet_name: Optional[str] = None,
                encoding: str = "utf-8", separator: str = ","):
    """Конвертація XLSX в CSV"""
    print(f"[INFO] Конвертація {input_xlsx} → {output_csv}")

    # Читання Excel
    if sheet_name:
        df = pd.read_excel(input_xlsx, sheet_name=sheet_name, dtype=str)
    else:
        df = pd.read_excel(input_xlsx, dtype=str)

    print(f"[INFO] Завантажено {len(df):,} рядків, {len(df.columns)} колонок")

    # Збереження в CSV
    df.to_csv(output_csv, index=False, encoding=encoding, sep=separator)

    print(f"[OK] Створено: {output_csv}")


# ============================================================================
# АНАЛІЗ ТА ЗВЕДЕННЯ
# ============================================================================

def frequency_analysis(df: pd.DataFrame, column: str,
                      case: str = "upper", strip_ws: bool = True,
                      drop_empty: bool = True) -> pd.DataFrame:
    """Частотний аналіз колонки"""
    col = resolve_single_column(df, column)
    vals = normalize_series(df[col], case=case, strip=strip_ws)

    if drop_empty:
        vals = vals.replace({"": None}).dropna()

    result = (
        vals.value_counts(dropna=False)
            .rename_axis(col)
            .reset_index(name="КІЛЬКІСТЬ")
            .sort_values(["КІЛЬКІСТЬ", col], ascending=[False, True])
            .reset_index(drop=True)
    )

    return result


def unique_values(df: pd.DataFrame, column: str,
                 case: str = "upper", strip_ws: bool = True,
                 drop_empty: bool = True) -> pd.DataFrame:
    """Унікальні значення колонки"""
    col = resolve_single_column(df, column)
    vals = normalize_series(df[col], case=case, strip=strip_ws)

    if drop_empty:
        vals = vals.replace({"": None}).dropna()

    uniq = sorted(vals.unique(), key=lambda x: (x is None, str(x)))
    result = pd.DataFrame({col: uniq})

    return result


def deduplicate(df: pd.DataFrame, key_columns: List[str],
               normalize_keys: bool = True, keep: str = "first") -> pd.DataFrame:
    """Дедуплікація за ключовими колонками"""
    key_cols = [resolve_single_column(df, k) for k in key_columns]

    if normalize_keys:
        # Нормалізація ключів перед дедуплікацією
        temp_df = df.copy()
        for col in key_cols:
            temp_df[col] = normalize_series(temp_df[col], case="upper", strip=True)

        result = temp_df.drop_duplicates(subset=key_cols, keep=keep).copy()
    else:
        result = df.drop_duplicates(subset=key_cols, keep=keep).copy()

    print(f"[INFO] Дедуплікація: {len(df):,} → {len(result):,} рядків")

    return result


# ============================================================================
# ОБ'ЄДНАННЯ ФАЙЛІВ/АРКУШІВ
# ============================================================================

def merge_files(file_paths: List[str], output: str,
               deduplicate_keys: Optional[List[str]] = None,
               filters: Optional[List[Dict[str, Any]]] = None):
    """Об'єднання декількох CSV/XLSX файлів в один"""
    print(f"[INFO] Об'єднання {len(file_paths)} файлів...")

    dfs = []
    for path in file_paths:
        print(f"  Читання: {path}")
        df = read_file_auto(path)

        # Застосування фільтрів
        if filters:
            df = apply_filters(df, filters)

        dfs.append(df)

    # Об'єднання
    result = pd.concat(dfs, ignore_index=True)
    print(f"[INFO] Об'єднано: {len(result):,} рядків")

    # Дедуплікація якщо потрібно
    if deduplicate_keys:
        result = deduplicate(result, deduplicate_keys)

    # Збереження
    if output.endswith(".xlsx"):
        save_to_excel(result, output)
    else:
        result.to_csv(output, index=False, encoding="utf-8")

    print(f"[OK] Результат збережено: {output}")


def merge_sheets(input_xlsx: str, output_sheet: str = "MERGED",
                deduplicate_keys: Optional[List[str]] = None,
                filters: Optional[List[Dict[str, Any]]] = None,
                sheet_names: Optional[List[str]] = None):
    """Об'єднання аркушів Excel файлу з різними структурами"""
    print(f"[INFO] Об'єднання аркушів з {input_xlsx}")

    wb = load_workbook(input_xlsx, data_only=True)

    # Вибір аркушів
    if sheet_names:
        sheets_to_merge = [wb[name] for name in sheet_names if name in wb.sheetnames]
    else:
        sheets_to_merge = list(wb.worksheets)

    # Побудова UNION схеми (всі колонки з усіх аркушів)
    all_columns = []
    columns_map = {}

    for ws in sheets_to_merge:
        header = [cell.value for cell in ws[1]]
        norm_header = [norm_col_name(h) for h in header]

        for i, (orig, norm) in enumerate(zip(header, norm_header)):
            if norm and norm not in columns_map:
                columns_map[norm] = orig
                all_columns.append(orig)

    print(f"[INFO] Об'єднана схема: {len(all_columns)} колонок")

    # Читання та об'єднання даних
    dfs = []
    for ws in sheets_to_merge:
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0]
        df = df[1:]

        # Додавання відсутніх колонок
        for col in all_columns:
            if col not in df.columns:
                df[col] = None

        df = df[all_columns]  # Впорядкування колонок

        # Застосування фільтрів
        if filters:
            df = apply_filters(df, filters)

        dfs.append(df)

    result = pd.concat(dfs, ignore_index=True)
    print(f"[INFO] Об'єднано: {len(result):,} рядків")

    # Дедуплікація
    if deduplicate_keys:
        result = deduplicate(result, deduplicate_keys)

    # Додавання результату в книгу
    ws_out = wb.create_sheet(output_sheet)

    # Запис заголовків
    for i, col in enumerate(result.columns, 1):
        ws_out.cell(row=1, column=i, value=col)

    # Запис даних
    for r_idx, row in enumerate(result.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws_out.cell(row=r_idx, column=c_idx, value=value)

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions

    wb.save(input_xlsx)
    print(f"[OK] Додано аркуш '{output_sheet}' до {input_xlsx}")


# ============================================================================
# ІНТЕРАКТИВНИЙ РЕЖИМ
# ============================================================================

def interactive_menu():
    """Головне інтерактивне меню"""
    print("\n" + "="*70)
    print("  УНІВЕРСАЛЬНИЙ ОБРОБНИК CSV ТА EXCEL ФАЙЛІВ")
    print("="*70)
    print("""
Оберіть операцію:

1.  Конвертація CSV → XLSX
2.  Конвертація XLSX → CSV
3.  Фільтрація файлу (CSV/XLSX)
4.  Частотний аналіз (підрахунок значень)
5.  Унікальні значення
6.  Об'єднання файлів
7.  Об'єднання аркушів Excel
8.  Дедуплікація даних
9.  Інформація про файл
0.  Вихід
    """)

    choice = input("Ваш вибір [0-9]: ").strip()

    if choice == "1":
        op_csv_to_xlsx()
    elif choice == "2":
        op_xlsx_to_csv()
    elif choice == "3":
        op_filter_file()
    elif choice == "4":
        op_frequency_analysis()
    elif choice == "5":
        op_unique_values()
    elif choice == "6":
        op_merge_files()
    elif choice == "7":
        op_merge_sheets()
    elif choice == "8":
        op_deduplicate()
    elif choice == "9":
        op_file_info()
    elif choice == "0":
        print("До побачення!")
        sys.exit(0)
    else:
        print("[ERROR] Невірний вибір")


# Операції меню
def op_csv_to_xlsx():
    """Операція: CSV → XLSX"""
    input_file = input("Вхідний CSV файл: ").strip()
    output_file = input("Вихідний XLSX файл [auto]: ").strip()

    if not output_file:
        output_file = Path(input_file).with_suffix(".xlsx")

    csv_to_xlsx(input_file, output_file)


def op_xlsx_to_csv():
    """Операція: XLSX → CSV"""
    input_file = input("Вхідний XLSX файл: ").strip()

    # Показати доступні аркуші
    xls = pd.ExcelFile(input_file)
    print(f"Доступні аркуші: {', '.join(xls.sheet_names)}")

    sheet_name = input("Назва аркуша [перший]: ").strip() or None
    output_file = input("Вихідний CSV файл [auto]: ").strip()

    if not output_file:
        output_file = Path(input_file).with_suffix(".csv")

    xlsx_to_csv(input_file, output_file, sheet_name=sheet_name)


def op_filter_file():
    """Операція: фільтрація"""
    input_file = input("Вхідний файл (CSV/XLSX): ").strip()
    output_file = input("Вихідний файл: ").strip()

    df = read_file_auto(input_file)
    print(f"[INFO] Завантажено: {len(df):,} рядків, {len(df.columns)} колонок")
    print(f"Колонки: {', '.join(df.columns[:10])}{'...' if len(df.columns) > 10 else ''}")

    filters = prompt_filters()
    result = apply_filters(df, filters)

    print(f"[INFO] Результат: {len(result):,} рядків")

    if output_file.endswith(".xlsx"):
        save_to_excel(result, output_file)
    else:
        result.to_csv(output_file, index=False, encoding="utf-8")

    print(f"[OK] Збережено: {output_file}")


def op_frequency_analysis():
    """Операція: частотний аналіз"""
    input_file = input("Вхідний файл (CSV/XLSX): ").strip()

    df = read_file_auto(input_file)
    print(f"Колонки: {', '.join(df.columns)}")

    column = input("Колонка для аналізу: ").strip()

    result = frequency_analysis(df, column)

    print(f"\n{result.to_string(index=False)}")

    save = input("\nЗберегти результат? (y/n) [n]: ").strip().lower()
    if save == "y":
        output_file = input("Вихідний файл: ").strip()
        if output_file.endswith(".xlsx"):
            save_to_excel(result, output_file, sheet_name="Frequency")
        else:
            result.to_csv(output_file, index=False, encoding="utf-8")
        print(f"[OK] Збережено: {output_file}")


def op_unique_values():
    """Операція: унікальні значення"""
    input_file = input("Вхідний файл (CSV/XLSX): ").strip()

    df = read_file_auto(input_file)
    print(f"Колонки: {', '.join(df.columns)}")

    column = input("Колонка: ").strip()

    result = unique_values(df, column)

    print(f"\n[INFO] Знайдено {len(result)} унікальних значень")
    print(f"\n{result.head(20).to_string(index=False)}")

    if len(result) > 20:
        print(f"... (показано перші 20 з {len(result)})")

    save = input("\nЗберегти результат? (y/n) [n]: ").strip().lower()
    if save == "y":
        output_file = input("Вихідний файл: ").strip()
        if output_file.endswith(".xlsx"):
            save_to_excel(result, output_file, sheet_name="Unique")
        else:
            result.to_csv(output_file, index=False, encoding="utf-8")
        print(f"[OK] Збережено: {output_file}")


def op_merge_files():
    """Операція: об'єднання файлів"""
    files_str = input("Файли для об'єднання (через пробіл): ").strip()
    files = files_str.split()

    output_file = input("Вихідний файл: ").strip()

    dedup = input("Дедуплікація? (y/n) [n]: ").strip().lower()
    dedup_keys = None
    if dedup == "y":
        keys_str = input("Ключові колонки (через кому): ").strip()
        dedup_keys = [k.strip() for k in keys_str.split(",")]

    merge_files(files, output_file, deduplicate_keys=dedup_keys)


def op_merge_sheets():
    """Операція: об'єднання аркушів"""
    input_file = input("XLSX файл: ").strip()

    xls = pd.ExcelFile(input_file)
    print(f"Аркуші: {', '.join(xls.sheet_names)}")

    sheets_str = input("Аркуші для об'єднання (через кому, або 'all') [all]: ").strip() or "all"

    if sheets_str.lower() == "all":
        sheet_names = None
    else:
        sheet_names = [s.strip() for s in sheets_str.split(",")]

    output_sheet = input("Назва вихідного аркуша [MERGED]: ").strip() or "MERGED"

    dedup = input("Дедуплікація? (y/n) [n]: ").strip().lower()
    dedup_keys = None
    if dedup == "y":
        keys_str = input("Ключові колонки (через кому): ").strip()
        dedup_keys = [k.strip() for k in keys_str.split(",")]

    merge_sheets(input_file, output_sheet=output_sheet,
                deduplicate_keys=dedup_keys, sheet_names=sheet_names)


def op_deduplicate():
    """Операція: дедуплікація"""
    input_file = input("Вхідний файл (CSV/XLSX): ").strip()

    df = read_file_auto(input_file)
    print(f"Колонки: {', '.join(df.columns)}")

    keys_str = input("Ключові колонки для унікальності (через кому): ").strip()
    keys = [k.strip() for k in keys_str.split(",")]

    result = deduplicate(df, keys)

    output_file = input("Вихідний файл: ").strip()

    if output_file.endswith(".xlsx"):
        save_to_excel(result, output_file)
    else:
        result.to_csv(output_file, index=False, encoding="utf-8")

    print(f"[OK] Збережено: {output_file}")


def op_file_info():
    """Операція: інформація про файл"""
    input_file = input("Файл (CSV/XLSX): ").strip()

    p = Path(input_file)

    print(f"\n{'='*60}")
    print(f"Файл: {p.name}")
    print(f"Розмір: {p.stat().st_size / 1024 / 1024:.2f} МБ")
    print(f"{'='*60}")

    if p.suffix.lower() == ".csv":
        enc, sep = detect_encoding_and_sep(input_file)
        print(f"Кодування: {enc}")
        print(f"Роздільник: '{sep}'")

    df = read_file_auto(input_file)

    print(f"\nРядків: {len(df):,}")
    print(f"Колонок: {len(df.columns)}")
    print(f"\nКолонки:")
    for i, col in enumerate(df.columns, 1):
        print(f"  {i:3d}. {col}")

    print(f"\nПерші 5 рядків:")
    print(df.head().to_string(index=False))

    print(f"\nТипи даних:")
    print(df.dtypes.to_string())


# ============================================================================
# ГОЛОВНА ФУНКЦІЯ
# ============================================================================

def main():
    """Головна функція з CLI та інтерактивним режимом"""
    parser = argparse.ArgumentParser(
        description="Універсальний обробник CSV та Excel файлів",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Приклади використання:

  # Інтерактивний режим
  python data_processor.py

  # Конвертація CSV → XLSX
  python data_processor.py --csv-to-xlsx input.csv output.xlsx

  # Конвертація XLSX → CSV
  python data_processor.py --xlsx-to-csv input.xlsx output.csv

  # Об'єднання файлів
  python data_processor.py --merge file1.csv file2.csv -o merged.xlsx

  # Частотний аналіз
  python data_processor.py --frequency input.xlsx --column "Brand" -o result.xlsx
        """
    )

    parser.add_argument("--csv-to-xlsx", nargs=2, metavar=("INPUT", "OUTPUT"),
                       help="Конвертувати CSV в XLSX")
    parser.add_argument("--xlsx-to-csv", nargs=2, metavar=("INPUT", "OUTPUT"),
                       help="Конвертувати XLSX в CSV")
    parser.add_argument("--merge", nargs="+", metavar="FILE",
                       help="Об'єднати файли")
    parser.add_argument("--frequency", metavar="FILE",
                       help="Частотний аналіз")
    parser.add_argument("--column", help="Назва колонки")
    parser.add_argument("-o", "--output", help="Вихідний файл")
    parser.add_argument("--interactive", action="store_true",
                       help="Інтерактивний режим (за замовчуванням)")

    args = parser.parse_args()

    # CLI режим
    if args.csv_to_xlsx:
        csv_to_xlsx(args.csv_to_xlsx[0], args.csv_to_xlsx[1])

    elif args.xlsx_to_csv:
        xlsx_to_csv(args.xlsx_to_csv[0], args.xlsx_to_csv[1])

    elif args.merge:
        if not args.output:
            print("[ERROR] Вкажіть вихідний файл через -o")
            sys.exit(1)
        merge_files(args.merge, args.output)

    elif args.frequency:
        if not args.column:
            print("[ERROR] Вкажіть колонку через --column")
            sys.exit(1)
        df = read_file_auto(args.frequency)
        result = frequency_analysis(df, args.column)
        if args.output:
            if args.output.endswith(".xlsx"):
                save_to_excel(result, args.output, sheet_name="Frequency")
            else:
                result.to_csv(args.output, index=False, encoding="utf-8")
            print(f"[OK] Збережено: {args.output}")
        else:
            print(result.to_string(index=False))

    # Інтерактивний режим (за замовчуванням)
    else:
        while True:
            try:
                interactive_menu()
            except KeyboardInterrupt:
                print("\n\n[INFO] Перервано користувачем")
                sys.exit(0)
            except Exception as e:
                print(f"\n[ERROR] {e}")
                import traceback
                traceback.print_exc()
                input("\nНатисніть Enter для продовження...")


if __name__ == "__main__":
    main()
