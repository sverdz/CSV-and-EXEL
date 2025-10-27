#!/usr/bin/env python3

# -*- coding: utf-8 -*-
"""
xlsx_group_summary_interactive_v6.py
------------------------------------
Інтерактивне зведення «значення → кількість» або «унікальні значення» з універсальними фільтрами
та опціональною дедуплікацією рядків за ключовою(ими) колонкою(ами) перед агрегуванням.

Режими:
  1) Парні колонки (B або B:D або B,D,F чи назви — кожна колонка підсумовується окремо).
  2) Один стовпець (за назвою або літерою) — підрахунок частот або вивід унікальних значень.

Що нового у v6:
- «Один набір фільтрів для всіх аркушів» (як у v5) +
- НОВА опція: перед агрегуванням залишати лише унікальні рядки за заданою колонкою/колонками
  (дедуплікація через drop_duplicates), один раз задається й застосовується до всіх аркушів.
- Толерантний вибір колонок (назва або літера чи діапазони), нормалізація (keep/upper/lower),
  обрізання пробілів, прогрес-бари.

Залежності:  py -m pip install openpyxl pandas tqdm
Запуск:      py xlsx_group_summary_interactive_v6.py
"""
import sys
import re
import datetime as dt
from pathlib import Path
from typing import List, Tuple, Optional, Dict, Any

try:
    from tqdm import tqdm
except Exception:
    tqdm = None

import pandas as pd
import math
import numpy as np

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except Exception as e:
    print("[ERR] Потрібно встановити openpyxl: py -m pip install openpyxl")
    raise

FILTERS_HELP = """
[Довідка з фільтрів 0..6]
0) Без фільтрації — пропускає всі рядки.
1) Текст дорівнює — точний збіг; доступні нормалізація (keep/upper/lower) та обрізання пробілів.
2) Текст містить — пошук підрядка з нормалізацією/обрізанням.
3) Текст у списку — значення входить до переліку через ';' (наприклад: A;B;C).
4) Число дорівнює — числова рівність; нечислові значення відсіюються.
5) Число в діапазоні — відбір у межах [min; max], включно; кома/крапка як роздільник.
6) REGEX — регулярний вираз Python (для повного збігу використовуйте ^ і $).
Усі задані фільтри комбінуються через AND (логічне «І»).
"""

# ---------- утиліти ----------

def norm_name(s: str) -> str:
    if s is None:
        return ""
    t = str(s).strip().lower()
    t = t.replace("і", "i")
    cyr_to_lat = str.maketrans({
        "а":"a","в":"b","с":"c","е":"e","н":"h","о":"o","р":"p","к":"k","х":"x","т":"t","і":"i","м":"m"
    })
    t = t.translate(cyr_to_lat)
    t = re.sub(r"[\s\-\u00A0]+", "", t)
    return t

def resolve_single_column(df: pd.DataFrame, user_token: str) -> str:
    if user_token in df.columns:
        return user_token
    token = str(user_token).strip()
    if re.fullmatch(r"[A-Za-z]+", token):
        try:
            idx = column_index_from_string(token.upper()) - 1
            if 0 <= idx < df.shape[1]:
                return df.columns[idx]
        except Exception:
            pass
    target = norm_name(user_token)
    for col in df.columns:
        if norm_name(col) == target:
            return col
    for col in df.columns:
        if target in norm_name(col) or norm_name(col) in target:
            return col
    raise KeyError(f"Стовпця '{user_token}' не знайдено. Доступні: {list(df.columns)}")

def resolve_multi_columns(df: pd.DataFrame, token: str) -> List[str]:
    """
    Приймає рядок на кшталт:
      - B або B:D або B,D,F
      - Назви колонок через кому: Brand, Model
    Повертає список фактичних назв колонок у df.
    """
    tok = token.strip()
    cols: List[str] = []
    if re.fullmatch(r"[A-Za-z]+:[A-Za-z]+", tok):
        a, b = tok.split(":")
        ia = column_index_from_string(a.upper()) - 1
        ib = column_index_from_string(b.upper()) - 1
        if ia > ib:
            ia, ib = ib, ia
        for j in range(max(0, ia), min(df.shape[1]-1, ib) + 1):
            cols.append(df.columns[j])
        return cols
    if re.fullmatch(r"[A-Za-z]+(,[A-Za-z]+)*", tok):
        for t in tok.split(","):
            idx = column_index_from_string(t.strip().upper()) - 1
            cols.append(df.columns[idx])
        return cols
    # інакше — перелік назв
    names = [t.strip() for t in tok.split(",") if t.strip()]
    for nm in names:
        cols.append(resolve_single_column(df, nm))
    return cols

def parse_sheet_selection(all_names: List[str], user_input: str) -> List[str]:
    ui = user_input.strip()
    if ui.lower() in ("all", "всі", "все", "*", "ALL"):
        return all_names
    parts = [p.strip() for p in ui.split(",") if p.strip()]
    chosen = []
    for p in parts:
        if p.isdigit():
            idx = int(p)
            if 0 <= idx < len(all_names):
                chosen.append(all_names[idx])
            else:
                raise IndexError(f"Немає аркуша з індексом {idx}; доступні 0..{len(all_names)-1}")
        else:
            if p in all_names:
                chosen.append(p)
            else:
                raise KeyError(f"Немає аркуша з назвою '{p}'. Доступні: {all_names}")
    return chosen

def ask(prompt: str, default: Optional[str] = None) -> str:
    sfx = f" [{default}]" if default is not None else ""
    while True:
        val = input(f"{prompt}{sfx}: ").strip()
        if val == "" and default is not None:
            return default
        if val != "":
            return val

def yesno(prompt: str, default: bool = False) -> bool:
    d = "Y/n" if default else "y/N"
    while True:
        val = input(f"{prompt} ({d}): ").strip().lower()
        if val == "" and default is not None:
            return default
        if val in ("y","yes","т","так","+","1"):
            return True
        if val in ("n","no","ні","-","0"):
            return False

def normalize_series(s: pd.Series, case: str = "upper", strip: bool = True) -> pd.Series:
    s = s.fillna("").astype(str)
    if strip:
        s = s.str.strip()
    if case == "upper":
        s = s.str.upper()
    elif case == "lower":
        s = s.str.lower()
    return s

# ---------- фільтри: специфікація та застосування ----------

def prompt_filters_spec() -> List[Dict[str, Any]]:
    print(FILTERS_HELP)
    try:
        n = int(ask("Скільки фільтрів застосувати? (0..6)", default="0"))
    except Exception:
        n = 0
    n = max(0, min(6, n))
    specs: List[Dict[str, Any]] = []
    for i in range(1, n+1):
        print(f"\nФільтр {i}:")
        try:
            t = int(ask("Тип (1=TXT==, 2=TXT CONTAINS, 3=TXT IN LIST, 4=NUM==, 5=NUM RANGE, 6=REGEX)", default="2"))
        except Exception:
            t = 2
        spec: Dict[str, Any] = {"type": t}
        spec["col_token"] = ask("Стовпець (назва або літера)", default="A")
        if t in (1,2,3,6):
            spec["case"] = ask("Нормалізація для тексту (keep/upper/lower)", default="upper").lower()
            spec["strip_ws"] = yesno("Обрізати пробіли зліва/справа?", default=True)
        if t == 1:
            spec["eq_value"] = ask("Значення для рівності", default="")
        elif t == 2:
            spec["substr"] = ask("Підрядок для пошуку", default="")
        elif t == 3:
            raw = ask("Список значень (через ';')", default="A;B;C")
            spec["list_values"] = [x.strip() for x in raw.split(";") if x.strip()]
        elif t == 4:
            v = ask("Числове значення для рівності", default="0")
            try:
                spec["num_eq"] = float(v.replace(",", "."))
            except Exception:
                spec["num_eq"] = float("nan")
        elif t == 5:
            vmin = ask("Мінімум", default="0")
            vmax = ask("Максимум", default="9999999")
            try:
                spec["num_min"] = float(vmin.replace(",", "."))
                spec["num_max"] = float(vmax.replace(",", "."))
            except Exception:
                spec["num_min"], spec["num_max"] = float("-inf"), float("inf")
        elif t == 6:
            spec["pattern"] = ask("REGEX (Python re)", default=".*")
        specs.append(spec)
    return specs

def apply_filters_spec(df: pd.DataFrame, specs: List[Dict[str, Any]]) -> Optional[pd.Series]:
    if not specs:
        return None
    mask = pd.Series(True, index=df.index)
    for i, spec in enumerate(specs, 1):
        t = spec.get("type", 2)
        col_token = spec.get("col_token", "A")
        try:
            col = resolve_single_column(df, col_token)
        except Exception as e:
            print(f"[WARN] Фільтр {i} пропущено ({col_token}): {e}")
            continue

        if t in (1,2,3,6):
            case = spec.get("case", "upper")
            strip_ws = spec.get("strip_ws", True)
            s = normalize_series(df[col], case=case, strip=strip_ws)
        else:
            s = pd.to_numeric(df[col], errors="coerce")

        if t == 1:
            val = spec.get("eq_value", "")
            if strip_ws:
                val = val.strip()
            if case == "upper":
                val = val.upper()
            elif case == "lower":
                val = val.lower()
            mask &= (s == val)
        elif t == 2:
            sub = spec.get("substr", "")
            if case == "upper":
                sub = sub.upper()
            elif case == "lower":
                sub = sub.lower()
            mask &= s.str.contains(re.escape(sub), na=False)
        elif t == 3:
            vals = spec.get("list_values", [])
            if case == "upper":
                vals = [x.upper() for x in vals]
            elif case == "lower":
                vals = [x.lower() for x in vals]
            mask &= s.isin(vals)
        elif t == 4:
            v = spec.get("num_eq", float("nan"))
            mask &= (s == v)
        elif t == 5:
            vmin = spec.get("num_min", float("-inf"))
            vmax = spec.get("num_max", float("inf"))
            mask &= s.between(vmin, vmax, inclusive="both")
        elif t == 6:
            pattern = spec.get("pattern", ".*")
            try:
                rx = re.compile(pattern)
                mask &= s.map(lambda x: bool(rx.search(x)) if isinstance(x, str) else False)
            except Exception as e:
                print(f"[WARN] REGEX помилка у фільтрі {i}: {e}; фільтр пропущено")
        else:
            print(f"[WARN] Невідомий тип фільтра {t}; пропущено")
    return mask

# ---------- обчислення ----------

def frequency_one_column(df: pd.DataFrame, column_token: str,
                         case: str = "upper", strip_ws: bool = True,
                         drop_empty: bool = True) -> pd.DataFrame:
    col = resolve_single_column(df, column_token)
    vals = normalize_series(df[col], case=case, strip=strip_ws)
    if drop_empty:
        vals = vals.replace({"": None}).dropna()
    out = (
        vals.value_counts(dropna=False)
            .rename_axis(col)
            .reset_index(name="КІЛЬКІСТЬ")
            .sort_values(["КІЛЬКІСТЬ", col], ascending=[False, True], kind="mergesort")
            .reset_index(drop=True)
    )
    return out

def unique_values_one_column(df: pd.DataFrame, column_token: str,
                             case: str = "upper", strip_ws: bool = True,
                             drop_empty: bool = True) -> pd.DataFrame:
    col = resolve_single_column(df, column_token)
    vals = normalize_series(df[col], case=case, strip=strip_ws)
    if drop_empty:
        vals = vals.replace({"": None}).dropna()
    uniq = pd.DataFrame({col: sorted(vals.unique(), key=lambda x: (x is None, str(x))) })
    uniq["КІЛЬКІСТЬ"] = pd.NA
    return uniq

def to_excel_value(x):
    """Convert Pandas/NumPy NA/NaN and numpy scalars to Excel-friendly values.
    - Missing values (pd.NA, np.nan, None) -> None (blank cell)
    - NumPy scalars -> Python native via .item()
    - float('nan') -> None
    """
    # Fast path for None
    if x is None:
        return None
    # Handle pandas NA / numpy NaN
    try:
        import pandas as _pd
        if _pd.isna(x):
            return None
    except Exception:
        pass
    # Normalize numpy scalar types to Python
    try:
        import numpy as _np
        if isinstance(x, (_np.generic,)):
            x = x.item()
        # If after conversion it's a float nan, blank it
        if isinstance(x, float):
            if x != x:  # NaN check
                return None
    except Exception:
        # If numpy isn't available for some reason, ignore and return x
        pass
    return x

# ---------- запис ----------

def write_summary_sheet(wb, sheet_name: str, pairs: List[Tuple[str, pd.DataFrame]]):
    if sheet_name in wb.sheetnames:
        base = sheet_name
        i = 2
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base} ({i})"
            i += 1
    ws = wb.create_sheet(sheet_name)
    row = 1
    iterable = pairs
    if tqdm:
        iterable = tqdm(pairs, desc="Запис блоків", unit="блок")
    for block_name, df in iterable:
        ws.cell(row=row, column=1, value=block_name); row += 1
        ws.cell(row=row, column=1, value=df.columns[0])
        ws.cell(row=row, column=2, value=df.columns[1]); row += 1
        data_iter = df.itertuples(index=False, name=None)
        if tqdm:
            data_iter = tqdm(list(data_iter), desc=f"→ {block_name}", unit="ряд", leave=False)
        for val, cnt in data_iter:
            ws.cell(row=row, column=1, value=to_excel_value(val))
            ws.cell(row=row, column=2, value=to_excel_value(cnt))
            row += 1
        row += 1
    for col_idx in (1, 2):
        max_len = 0
        for r in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            v = r[0].value
            lv = len(str(v)) if v is not None else 0
            if lv > max_len:
                max_len = lv
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)
    ws.freeze_panes = "A2"
    first_header_row = None
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=2).value == "КІЛЬКІСТЬ":
            first_header_row = i
            break
    if first_header_row:
        end = first_header_row + 1
        while end <= ws.max_row and ws.cell(row=end, column=1).value not in (None, ""):
            end += 1
        ws.auto_filter.ref = f"A{first_header_row}:B{max(first_header_row, end-1)}"

# ---------- головний інтерфейс ----------

def main():
    print("== Інтерактивне зведення v6 ==")
    path_str = ask("Шлях до файла (XLSX/CSV)", default="input.xlsx")
    src = Path(path_str)
    if not src.exists():
        print(f"[ERR] Файл не знайдено: {src}")
        sys.exit(1)

    mode = ask("Режим: 1) парні колонки  2) один стовпець", default="2")
    try:
        mode = int(mode)
    except Exception:
        mode = 2
    if mode not in (1,2):
        mode = 2

    if src.suffix.lower() == ".csv":
        print("[INFO] CSV — працюємо з одним «аркушем».")
        df_csv = pd.read_csv(src, dtype=str)
        all_names = ["CSV"]
        sheets_reader = {"CSV": df_csv}
    else:
        xls = pd.ExcelFile(src)
        all_names = xls.sheet_names
        sheets_reader = None

    print("Доступні аркуші:")
    for i, n in enumerate(all_names):
        print(f"  [{i}] {n}")
    sel = ask("Вибір аркушів (all або індекси/назви через кому)", default="all")
    try:
        chosen = parse_sheet_selection(all_names, sel)
    except Exception as e:
        print(f"[ERR] {e}")
        sys.exit(2)

    case = ask("Нормалізація значень (keep/upper/lower)", default="upper").lower()
    if case not in ("keep","upper","lower"):
        case = "upper"
    strip_ws = yesno("Обрізати пробіли зліва/справа?", default=True)
    output_kind = ask("Вивід: 1) частоти  2) унікальні значення", default="1")
    output_kind = "unique" if str(output_kind).strip() == "2" else "freq"

    # Фільтри (глобально один раз або поаркушно)
    use_filters = yesno("Застосувати фільтри перед агрегуванням?", default=False)
    global_spec: Optional[List[Dict[str, Any]]] = None
    if use_filters:
        one_for_all = yesno("Задати фільтри ОДИН РАЗ для всіх аркушів?", default=True)
        if one_for_all:
            global_spec = prompt_filters_spec()

    # Дедуплікація
    do_dedup = yesno("Перед агрегуванням залишати лише УНІКАЛЬНІ рядки за колонкою/колонками?", default=False)
    dedup_tokens: Optional[str] = None
    if do_dedup:
        hint = " (можна: A або A:C або A,C,E або назви через кому)"
        if mode == 2:
            # за замовчуванням — цільова колонка для агрегування
            default_key = "A"
            dedup_tokens = ask("Ключ для унікальності" + hint, default=default_key)
        else:
            dedup_tokens = ask("Ключ для унікальності" + hint, default="A")

    pairs: List[Tuple[str, pd.DataFrame]] = []
    overall_list = []

    if mode == 2:
        column_token = ask("Колонка (назва або літера)", default="A")
        iterable = chosen
        if tqdm:
            iterable = tqdm(chosen, desc="Обробка аркушів", unit="аркуш")
        for name in iterable:
            if src.suffix.lower() == ".csv":
                df = sheets_reader["CSV"].copy()
            else:
                df = pd.read_excel(src, sheet_name=name, dtype=str)

            # фільтри
            if use_filters:
                if global_spec is not None:
                    mask = apply_filters_spec(df, global_spec)
                else:
                    print(FILTERS_HELP)
                    print(f"[ФІЛЬТРИ] Аркуш: {name}")
                    spec = prompt_filters_spec()
                    mask = apply_filters_spec(df, spec)
                if mask is not None:
                    df = df[mask].copy()

            # дедуплікація
            if do_dedup and dedup_tokens:
                try:
                    key_cols = resolve_multi_columns(df, dedup_tokens)
                    df = df.drop_duplicates(subset=key_cols, keep="first").copy()
                except Exception as e:
                    print(f"[WARN] Дедуплікація пропущена для '{name}': {e}")

            # агрегування
            try:
                if output_kind == "freq":
                    out = frequency_one_column(df, column_token, case=case, strip_ws=strip_ws, drop_empty=True)
                else:
                    out = unique_values_one_column(df, column_token, case=case, strip_ws=strip_ws, drop_empty=True)
            except Exception as e:
                print(f"[WARN] Пропущено '{name}': {e}")
                continue
            pairs.append((f"{name} — {column_token}", out))
            overall_list.append(out.rename(columns={out.columns[0]: "ЗНАЧЕННЯ"})[["ЗНАЧЕННЯ"]])

        if overall_list:
            concat_vals = pd.concat(overall_list, ignore_index=True)
            if output_kind == "freq":
                grp = (concat_vals.value_counts().rename_axis("ЗНАЧЕННЯ").reset_index(name="КІЛЬКІСТЬ")
                       .sort_values(["КІЛЬКІСТЬ","ЗНАЧЕННЯ"], ascending=[False, True], kind="mergesort").reset_index(drop=True))
                pairs.append(("ЗАГАЛОМ — підсумок", grp.rename(columns={"ЗНАЧЕННЯ": column_token})))
            else:
                uniq = pd.DataFrame({column_token: sorted(concat_vals["ЗНАЧЕННЯ"].unique(), key=lambda x: (x is None, str(x)))})
                uniq["КІЛЬКІСТЬ"] = pd.NA
                pairs.append(("ЗАГАЛОМ — унікальні", uniq))

    else:
        # парні колонки: кожну обрану колонку агрегуємо окремо
        cols_token = ask("Колонки (B або B:D або B,D,F або назви)", default="B")

        def token_to_indices(df: pd.DataFrame, token: str) -> List[int]:
            token = token.strip()
            if re.fullmatch(r"[A-Za-z]+:[A-Za-z]+", token):
                a, b = token.split(":")
                ia = column_index_from_string(a.upper()) - 1
                ib = column_index_from_string(b.upper()) - 1
                if ia > ib:
                    ia, ib = ib, ia
                return list(range(max(0,ia), min(df.shape[1]-1, ib)+1))
            if re.fullmatch(r"[A-Za-z]+(,[A-Za-z]+)*", token):
                return [column_index_from_string(t.strip().upper()) - 1 for t in token.split(",")]
            if re.fullmatch(r"[A-Za-z]+", token):
                return [column_index_from_string(token.upper()) - 1]
            names = [t.strip() for t in token.split(",")]
            idxs = []
            for nm in names:
                try:
                    col = resolve_single_column(df, nm)
                    idxs.append(df.columns.get_loc(col))
                except Exception:
                    pass
            if not idxs:
                raise ValueError("Не вдалося інтерпретувати перелік колонок.")
            return idxs

        iterable = chosen
        if tqdm:
            iterable = tqdm(chosen, desc="Обробка аркушів", unit="аркуш")
        for name in iterable:
            if src.suffix.lower() == ".csv":
                df = sheets_reader["CSV"].copy()
            else:
                df = pd.read_excel(src, sheet_name=name, dtype=str)

            # фільтри
            if use_filters:
                if global_spec is not None:
                    mask = apply_filters_spec(df, global_spec)
                else:
                    print(FILTERS_HELP)
                    print(f"[ФІЛЬТРИ] Аркуш: {name}")
                    spec = prompt_filters_spec()
                    mask = apply_filters_spec(df, spec)
                if mask is not None:
                    df = df[mask].copy()

            # дедуплікація
            if do_dedup and dedup_tokens:
                try:
                    key_cols = resolve_multi_columns(df, dedup_tokens)
                    df = df.drop_duplicates(subset=key_cols, keep="first").copy()
                except Exception as e:
                    print(f"[WARN] Дедуплікація пропущена для '{name}': {e}")

            # агрегування
            try:
                idxs = token_to_indices(df, cols_token)
            except Exception as e:
                print(f"[WARN] Пропущено '{name}': {e}")
                continue

            iter_cols = idxs
            if tqdm:
                iter_cols = tqdm(idxs, desc=f"Колонки {name}", unit="col", leave=False)
            for idx in iter_cols:
                col = df.columns[idx]
                if output_kind == "freq":
                    out = frequency_one_column(df, col, case=case, strip_ws=strip_ws, drop_empty=True)
                else:
                    out = unique_values_one_column(df, col, case=case, strip_ws=strip_ws, drop_empty=True)
                pairs.append((f"{name} — {col}", out))

    if not pairs:
        print("[ERR] Немає даних для запису.")
        sys.exit(3)

    same = yesno("Додати SUMMARY у цей самий XLSX? (для CSV буде окремий файл)", default=True)
    if src.suffix.lower() == ".csv" or not same:
        out_path = Path(f"summary_{dt.date.today().isoformat()}.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        write_summary_sheet(wb, "SUMMARY", pairs)
        wb.save(out_path)
        print(f"[OK] Збережено окремий файл: {out_path.resolve()}")
    else:
        wb = load_workbook(src)
        write_summary_sheet(wb, "SUMMARY", pairs)
        wb.save(src)
        print(f"[OK] Додано аркуш 'SUMMARY' до: {src.resolve()}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[INTERRUPTED] Перервано користувачем.")
        sys.exit(130)
