# xlsx_unify_unique_interactive_v2.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ІНТЕРАКТИВНЕ злиття аркушів із різними шапками:
- Будує «об’єднану» шапку як UNION усіх колонок з усіх аркушів-джерел.
- Автозіставлення назв колонок без урахування регістру/пробілів/дефісів + мапа кириличних «двійників».
- Відсутні на конкретному аркуші колонки заповнюються порожнім.
- Фільтр 1..6, дедуплікація за 1..N ключами, нормалізація ключів.
- Пише результат у ТУ Ж саму книгу на новий аркуш (типово UNIQUE).

Запуск: python xlsx_unify_unique_interactive_v2.py
"""

import sys, re
from pathlib import Path
from typing import List, Tuple, Set, Optional, Callable, Dict, Any

try:
    from openpyxl import load_workbook
except Exception:
    sys.exit("Не знайдено openpyxl. Встановіть:  py -m pip install openpyxl")

# -------- утиліти вводу/виводу --------

def ask(prompt: str, default: Optional[str] = None) -> str:
    s = input(f"{prompt}" + (f" [{default}]" if default is not None else "") + ": ").strip()
    return (default if (not s and default is not None) else s)

def yesno(prompt: str, default_yes: bool = True) -> bool:
    tag = "[Y/n]" if default_yes else "[y/N]"
    s = input(f"{prompt} {tag}: ").strip().lower()
    if not s:
        return default_yes
    return s.startswith("y")

def pick_output_sheet_name(existing: List[str], base: str) -> str:
    if base not in existing:
        return base
    i = 2
    while True:
        cand = f"{base} ({i})"
        if cand not in existing:
            return cand
        i += 1

# -------- нормалізація назв колонок і значень --------

CYR_TO_LAT = str.maketrans({
    "А":"A","В":"B","С":"C","Е":"E","Н":"H","К":"K","М":"M","О":"O","Р":"P","Т":"T","Х":"X","У":"Y","І":"I","Ї":"I","Й":"I","Ґ":"G",
    "а":"a","в":"b","с":"c","е":"e","н":"h","к":"k","м":"m","о":"o","р":"p","т":"t","х":"x","у":"y","і":"i","ї":"i","й":"i","ґ":"g"
})

def norm_col_name(name: Any) -> str:
    s = "" if name is None else str(name)
    s = s.translate(CYR_TO_LAT)           # латинізація схожих букв
    s = s.strip().lower()
    s = re.sub(r"[\s\-]+", "_", s)        # пробіли/дефіси → _
    s = re.sub(r"[^0-9a-z_]+", "", s)     # прибрати інше
    s = re.sub(r"_+", "_", s).strip("_")
    return s

def canon_header(ws, header_row: int) -> List[str]:
    return [("" if c.value is None else str(c.value)) for c in ws[header_row]]

def normalize_key(values: Tuple, upper: bool, strip_spaces: bool, drop_spaces: bool, drop_dashes: bool) -> Tuple:
    out = []
    for v in values:
        s = "" if v is None else str(v)
        if strip_spaces: s = s.strip()
        if drop_spaces:  s = s.replace(" ", "")
        if drop_dashes:  s = s.replace("-", "")
        if upper:        s = s.upper()
        out.append(s)
    return tuple(out)

# -------- фільтр 1..6 --------

def build_filter_fn(header_display: List[str], spec: Dict) -> Callable[[Tuple], bool]:
    mode = spec.get("mode", "6")
    raw_col = spec.get("column", "")
    low = {h.strip().lower(): i for i, h in enumerate(header_display)}
    col_idx = low.get(raw_col.strip().lower(), None)

    if mode == "6" or col_idx is None:
        return lambda row: True

    if mode == "1":
        val = spec.get("value", "")
        return lambda row: ("" if row[col_idx] is None else str(row[col_idx])) == val

    if mode == "2":
        sub = str(spec.get("value", "")).lower()
        return lambda row: sub in ("" if row[col_idx] is None else str(row[col_idx]).lower())

    if mode == "3":
        values = set(str(x).lower() for x in spec.get("values", []))
        return lambda row: ("" if row[col_idx] is None else str(row[col_idx]).lower()) in values

    if mode == "4":
        try: target = float(str(spec.get("value","")).replace(",", "."))
        except Exception: target = None
        def fn(row):
            if target is None: return False
            try: x = float(str(row[col_idx]).replace(",", "."))
            except Exception: return False
            return x == target
        return fn

    if mode == "5":
        def tofloat(x):
            try: return float(str(x).replace(",", "."))
            except Exception: return None
        vmin = tofloat(spec.get("min","")); vmax = tofloat(spec.get("max",""))
        def fn(row):
            x = tofloat(row[col_idx])
            if x is None: return False
            if vmin is not None and x < vmin: return False
            if vmax is not None and x > vmax: return False
            return True
        return fn

    return lambda row: True

def prompt_filter_params_interactive() -> Dict:
    print("\n=== Налаштування фільтра ===")
    col = ask("Назва колонки для фільтра (як у заголовку; порожньо = без фільтра)", default="")
    if not col:
        print("Фільтр вимкнено.")
        return {"mode":"6","column":""}
    print("Тип значення:\n 1) Текст ==\n 2) Містить (icase)\n 3) Перелік (isin)\n 4) Число ==\n 5) Число в діапазоні [min;max]\n 6) Без фільтра")
    mode = ask("Оберіть 1/2/3/4/5/6", default="6")
    spec = {"column": col, "mode": mode, "force_text": False}
    if mode == "1":
        spec["value"] = ask("Значення (текст)")
        spec["force_text"] = yesno("Форматувати цю колонку у вихідному аркуші як ТЕКСТ?", default_yes=True)
    elif mode == "2":
        spec["value"] = ask("Підрядок (текст)")
        spec["force_text"] = yesno("Форматувати цю колонку як ТЕКСТ?", default_yes=True)
    elif mode == "3":
        vals = ask("Перелік значень через кому")
        spec["values"] = [v.strip() for v in vals.split(",") if v.strip()]
        spec["force_text"] = yesno("Форматувати цю колонку як ТЕКСТ?", default_yes=True)
    elif mode == "4":
        spec["value"] = ask("Числове значення (==)")
    elif mode == "5":
        spec["min"] = ask("Мінімум"); spec["max"] = ask("Максимум")
    else:
        print("Фільтр вимкнено.")
    return spec

# -------- побудова UNION-схеми --------

def build_union_schema(sheets, header_row: int, base_out_name: str) -> Tuple[List[str], List[str], Dict[str, Dict[int,int]]]:
    """
    Повертає:
      display_header: «людські» назви колонок (перше зустрічне ім'я для кожного нормалізованого ключа),
      norm_header   : нормалізовані ключі,
      mapping_per_ws: для кожного аркуша — {canonical_col_idx -> ws_col_idx або -1, якщо колонки нема}
    """
    skip_name = base_out_name.lower()
    data_sheets = [ws for ws in sheets if ws.title.lower() != skip_name and not ws.title.lower().startswith(f"{skip_name} (")]

    order: List[str] = []
    display_for_norm: Dict[str,str] = {}
    headers_raw: Dict[str, List[str]] = {}
    headers_norm: Dict[str, List[str]] = {}

    for ws in data_sheets:
        hdr = canon_header(ws, header_row)
        headers_raw[ws.title] = hdr
        norms = [norm_col_name(x) for x in hdr]
        headers_norm[ws.title] = norms
        for i, nk in enumerate(norms):
            if nk == "":
                continue
            if nk not in display_for_norm:
                display_for_norm[nk] = hdr[i]
                order.append(nk)

    norm_header = order[:]  # порядок першої появи
    display_header = [display_for_norm[nk] for nk in norm_header]

    mapping_per_ws: Dict[str, Dict[int,int]] = {}
    for ws in data_sheets:
        norms = headers_norm[ws.title]
        first_pos: Dict[str,int] = {}
        for j, nk in enumerate(norms):
            if nk and nk not in first_pos:
                first_pos[nk] = j
        mapping = {}
        for ci, nk in enumerate(norm_header):
            mapping[ci] = first_pos.get(nk, -1)
        mapping_per_ws[ws.title] = mapping

    return display_header, norm_header, mapping_per_ws

# -------- основний сценарій --------

def main():
    print("=== Об’єднання аркушів Excel з різними шапками у єдиний аркуш (унікальні записи) ===")
    here = Path.cwd()
    xlsx_path = Path(ask("Шлях до Excel-файлу (*.xlsx)", default=str(here/"out.xlsx")))
    # відлов можливого блокування файлу
    from time import sleep
    for _ in range(30):
        try:
            wb = load_workbook(xlsx_path, data_only=True)
            break
        except PermissionError:
            input(f"Файл заблоковано (Excel/OneDrive): {xlsx_path}\nЗакрийте і натисніть Enter…")
    else:
        raise

    base_out = ask("Назва вихідного аркуша", default="UNIQUE")
    out_name = pick_output_sheet_name(wb.sheetnames, base_out)
    ws_out = wb.create_sheet(out_name)

    print("\nВведіть ключові колонки (через кому), за якими рядки унікальні. Напр.: N_REG_NEW або VIN або REG_ADDR_KOATUU,OPER_COI")
    keys_raw = ask("Ключові колонки")
    user_keys = [s.strip() for s in keys_raw.split(",") if s.strip()]
    if not user_keys:
        wb.remove(ws_out); wb.save(xlsx_path); sys.exit("Не вказано ключові колонки.")

    to_upper     = yesno("Переводити ключові значення у ВЕРХНІЙ регістр?", default_yes=True)
    drop_spaces  = yesno("Видаляти ВСІ пробіли з ключових значень?", default_yes=False)
    drop_dashes  = yesno("Видаляти дефіси '-' з ключових значень?", default_yes=False)
    strip_edges  = True

    filt_spec = prompt_filter_params_interactive()

    # аркуші-джерела (пропускаємо підсумкові базової назви)
    skip_prefix = base_out.lower()
    data_sheets = [ws for ws in wb.worksheets if ws.title.lower() != skip_prefix and not ws.title.lower().startswith(f"{skip_prefix} (")]
    if not data_sheets:
        wb.remove(ws_out); wb.save(xlsx_path); sys.exit("Немає аркушів для злиття.")

    header_row = 1

    # UNION-схема і відповідності колонок
    display_header, norm_header, mapping_per_ws = build_union_schema(data_sheets, header_row, base_out)
    ws_out.append(display_header)

    # індекси ключових колонок у канонічній шапці
    display2idx = {h.strip().lower(): i for i, h in enumerate(display_header)}
    missing_keys = [k for k in user_keys if k.strip().lower() not in display2idx]
    if missing_keys:
        wb.remove(ws_out); wb.save(xlsx_path)
        sys.exit(f"Не знайдено ключових колонок у об’єднаній шапці: {missing_keys}\nДоступні: {display_header}")
    key_idx = [display2idx[k.strip().lower()] for k in user_keys]

    # текстовий формат для фільтрової колонки у виході (за потреби)
    force_text_col_index = None
    if filt_spec.get("mode") in ("1","2","3") and filt_spec.get("force_text", False):
        low = {h.strip().lower(): i for i, h in enumerate(display_header)}
        force_text_col_index = low.get(filt_spec.get("column","").strip().lower(), None)

    row_filter = build_filter_fn(display_header, filt_spec)

    # основний прохід
    seen: Set[Tuple] = set()
    total_rows = 0
    written_rows = 0

    for ws in data_sheets:
        mapping = mapping_per_ws[ws.title]  # canonical idx -> ws idx (або -1)
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            total_rows += 1

            # сформувати ряд у канонічному порядку
            out_row = []
            for ci in range(len(display_header)):
                j = mapping[ci]
                v = (row[j] if (j is not None and j >= 0 and j < len(row)) else None)
                out_row.append(v)

            # фільтр
            if not row_filter(tuple(out_row)):
                continue

            # ключ і унікальність
            key_values = tuple(out_row[i] for i in key_idx)
            norm_key = normalize_key(key_values, to_upper, strip_edges, drop_spaces, drop_dashes)
            if norm_key in seen:
                continue
            seen.add(norm_key)

            # ключові колонки → текст (щоб Excel не робив E+)
            for i in key_idx:
                out_row[i] = "" if out_row[i] is None else str(out_row[i])

            ws_out.append(out_row)
            written_rows += 1

            if force_text_col_index is not None:
                r = ws_out.max_row
                c = force_text_col_index + 1
                ws_out.cell(row=r, column=c).number_format = "@"

            if written_rows % 100000 == 0:
                print(f"[{ws.title}] processed={total_rows:,} unique_written={written_rows:,}")

        print(f"[{ws.title}] завершено: розглянуто {total_rows:,}, записано унікальних {written_rows:,}")

    ws_out.freeze_panes = "A2"
    try:
        ws_out.auto_filter.ref = ws_out.dimensions
    except Exception:
        pass

    wb.save(xlsx_path)
    print(f"\n[OK] Додано аркуш '{out_name}' у файл: {xlsx_path}")
    print(f"Підсумок: рядків розглянуто {total_rows:,}, унікальних записано {written_rows:,}")
    print("Примітка: об’єднана шапка включає всі колонки, що траплялися; відсутні на окремих аркушах заповнено порожнім.")

if __name__ == "__main__":
    main()
